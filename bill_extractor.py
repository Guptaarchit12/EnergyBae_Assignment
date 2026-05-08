"""
bill_extractor.py
─────────────────
Uses the Google Gemini API to extract structured data from
an MSEDCL / Maharashtra electricity bill (PDF or image),
then fills the Solar Load Calculator Excel template.

Usage:
    python src/bill_extractor.py --bill path/to/bill.pdf --output output/filled_report.xlsx
    python src/bill_extractor.py --bill path/to/bill.jpg  --output output/filled_report.xlsx
"""

import os
import sys
import json
import base64
import argparse
import shutil
from pathlib import Path
from datetime import datetime

from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ── Internal imports ──────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))
from create_template import create_template

# ── Constants ─────────────────────────────────────────────────────────────────
TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "solar_load_calculator_template.xlsx"

# Yellow for input cells
YELLOW_INPUT = "FFF9C4"

EXTRACTION_SCHEMA = {
    "consumer_name":          {"type": "string", "desc": "Customer/consumer name on the bill"},
    "consumer_number":        {"type": "string", "desc": "Consumer account/service number"},
    "address":                {"type": "string", "desc": "Consumer address"},
    "tariff_category":        {"type": "string", "desc": "Tariff category e.g. LT-I, LT-II, HT-I"},
    "division":               {"type": "string", "desc": "MSEDCL division or sub-division"},
    "bill_month":             {"type": "string", "desc": "Billing month/period e.g. March 2024"},
    "bill_date":              {"type": "string", "desc": "Bill issue date"},
    "meter_number":           {"type": "string", "desc": "Meter serial number"},
    "sanctioned_load_kw":     {"type": "number", "desc": "Sanctioned / contracted load in kW"},
    "units_consumed":         {"type": "number", "desc": "Total units (kWh) consumed in this billing period"},
    "bill_amount":            {"type": "number", "desc": "Total bill amount before any payment"},
    "fixed_charges":          {"type": "number", "desc": "Fixed or demand charges in ₹"},
    "electricity_duty":       {"type": "number", "desc": "Electricity duty component in ₹"},
    "fuel_adjustment_charge": {"type": "number", "desc": "Fuel Adjustment Charge (FAC) in ₹"},
    "meter_rent":             {"type": "number", "desc": "Meter rent and other miscellaneous charges in ₹"},
    "subsidies_rebate":       {"type": "number", "desc": "Any subsidy or rebate applied in ₹"},
    "net_payable":            {"type": "number", "desc": "Final net payable amount in ₹"},
    "tariff_slab":            {"type": "string", "desc": "Applicable tariff slab designation"},
    "rate_per_unit":          {"type": "number", "desc": "Effective or blended rate per unit (₹/kWh)"},
    "connected_load_kw":      {"type": "number", "desc": "Connected load in kW"},
    "power_factor":           {"type": "number", "desc": "Power factor if mentioned"},
    "monthly_consumption": {
        "type": "object",
        "desc": "Monthly consumption data if available (12-month history)",
        "keys": ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    }
}

SYSTEM_PROMPT = """You are an expert OCR and data extraction system specialised in Indian electricity bills,
particularly MSEDCL (Maharashtra State Electricity Distribution Co. Ltd) bills.

Your job is to read the bill content provided and extract structured data with high accuracy.
Return ONLY valid JSON — no markdown, no explanation, no preamble.

Rules:
1. Extract every field you can find. If a field is not present, use null.
2. All monetary values must be numbers (not strings), in Indian Rupees.
3. Units consumed must be a number (kWh).
4. For monthly_consumption, provide a dictionary with month abbreviations as keys and kWh values as numbers.
   Only include months that are explicitly listed in the bill.
5. If there is a 12-month consumption graph or table, extract all values.
6. Power factor: extract as decimal (e.g., 0.95 not 95%).
7. Rate per unit: calculate effective rate = net_payable / units_consumed if not explicitly stated.
8. Be conservative — do not guess values. Use null if uncertain.

Return exactly this JSON structure:
{
  "consumer_name": "...",
  "consumer_number": "...",
  "address": "...",
  "tariff_category": "...",
  "division": "...",
  "bill_month": "...",
  "bill_date": "...",
  "meter_number": "...",
  "sanctioned_load_kw": null,
  "units_consumed": null,
  "bill_amount": null,
  "fixed_charges": null,
  "electricity_duty": null,
  "fuel_adjustment_charge": null,
  "meter_rent": null,
  "subsidies_rebate": null,
  "net_payable": null,
  "tariff_slab": "...",
  "rate_per_unit": null,
  "connected_load_kw": null,
  "power_factor": null,
  "monthly_consumption": {}
}"""


def encode_file_to_base64(file_path: str) -> tuple[str, str]:
    """Returns (base64_data, media_type)."""
    path = Path(file_path)
    ext = path.suffix.lower()

    media_type_map = {
        ".pdf":  "application/pdf",
        ".png":  "image/png",
        ".jpg":  "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
        ".gif":  "image/gif",
    }
    media_type = media_type_map.get(ext)
    if not media_type:
        raise ValueError(f"Unsupported file type: {ext}. Supported: PDF, PNG, JPG, WEBP, GIF")

    with open(file_path, "rb") as f:
        data = base64.standard_b64encode(f.read()).decode("utf-8")
    return data, media_type


def extract_bill_data(bill_path: str, api_key: str | None = None) -> dict:
    """
    Calls Google Gemini with the bill file and returns extracted data dict.
    """
    api_key = api_key or os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        raise EnvironmentError(
            "GOOGLE_API_KEY not found. Set it as an environment variable or pass via --api-key."
        )

    client = genai.Client(api_key=api_key)
    b64_data, media_type = encode_file_to_base64(bill_path)

    print(f"  -> Sending bill to Gemini AI for extraction...")

    # Inline base64 data part (works for both PDFs and images)
    file_part = types.Part.from_bytes(
        data=base64.b64decode(b64_data),
        mime_type=media_type,
    )

    user_text = (
        "Please extract all available data from this electricity bill "
        "and return it as JSON following the exact schema specified. "
        "Return ONLY the JSON object."
    )

    response = client.models.generate_content(
        model="gemini-2.0-flash",
        contents=[file_part, user_text],
        config=types.GenerateContentConfig(
            system_instruction=SYSTEM_PROMPT,
            max_output_tokens=2000,
            temperature=0.0,
        ),
    )

    raw_text = response.text.strip()

    # Strip markdown code fences if present
    if raw_text.startswith("```"):
        raw_text = raw_text.split("```")[1]
        if raw_text.startswith("json"):
            raw_text = raw_text[4:]
        raw_text = raw_text.strip()

    data = json.loads(raw_text)
    print(f"  [OK] Extraction complete. Fields found: {sum(1 for v in data.values() if v is not None and v != {})}")
    return data


def fill_excel_template(extracted_data: dict, output_path: str,
                        template_path: str | None = None) -> str:
    """
    Fills the Solar Load Calculator Excel template with extracted bill data.
    Returns the path to the filled Excel file.
    """
    tpl = template_path or str(TEMPLATE_PATH)
    if not Path(tpl).exists():
        print("  → Template not found, generating fresh template...")
        create_template(tpl)

    # Copy template to output location
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(tpl, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Bill Input"]

    def fill(cell_ref: str, value, num_format: str | None = None):
        if value is None:
            return
        c = ws[cell_ref]
        c.value = value
        if num_format:
            c.number_format = num_format
        # Keep yellow fill for visual clarity
        c.fill = PatternFill("solid", start_color=YELLOW_INPUT)

    d = extracted_data

    # ── Section A: Customer Info ──────────────────────────────────────────────
    fill("B5", d.get("consumer_name"))
    fill("D5", d.get("consumer_number"))
    fill("B6", d.get("address"))
    fill("D6", d.get("tariff_category"))
    fill("B7", d.get("discom", "MSEDCL"))
    fill("D7", d.get("division"))
    fill("B8", d.get("bill_month"))
    fill("D8", d.get("bill_date"))
    fill("B9", d.get("meter_number"))
    fill("D9", d.get("sanctioned_load_kw"), "#,##0.0")

    # ── Section B: Monthly Consumption ───────────────────────────────────────
    monthly = d.get("monthly_consumption") or {}
    month_cols = {
        "Jan": 2, "Feb": 3, "Mar": 4, "Apr": 5,  "May": 6,  "Jun": 7,
        "Jul": 8, "Aug": 9, "Sep": 10, "Oct": 11, "Nov": 12, "Dec": 13
    }
    for month, col in month_cols.items():
        val = monthly.get(month)
        if val is not None:
            c = ws.cell(row=13, column=col)
            c.value = float(val)
            c.number_format = "#,##0.0"
            c.fill = PatternFill("solid", start_color=YELLOW_INPUT)

    # If no monthly data but we have units_consumed for current month, fill that month
    if not monthly and d.get("units_consumed") and d.get("bill_month"):
        bm = str(d.get("bill_month", ""))
        for abbr in month_cols:
            if abbr.lower() in bm.lower():
                c = ws.cell(row=13, column=month_cols[abbr])
                c.value = float(d["units_consumed"])
                c.fill = PatternFill("solid", start_color=YELLOW_INPUT)
                break

    # ── Section C: Bill Details ───────────────────────────────────────────────
    fill("B17", d.get("units_consumed"),         "#,##0")
    fill("D17", d.get("bill_amount"),             "#,##0.00")
    fill("B18", d.get("fixed_charges"),           "#,##0.00")
    fill("D18", d.get("electricity_duty"),        "#,##0.00")
    fill("B19", d.get("fuel_adjustment_charge"),  "#,##0.00")
    fill("D19", d.get("meter_rent"),              "#,##0.00")
    fill("B20", d.get("subsidies_rebate"),        "#,##0.00")
    fill("D20", d.get("net_payable"),             "#,##0.00")
    fill("B21", d.get("tariff_slab"))
    fill("D21", d.get("rate_per_unit"),           "#,##0.00")
    fill("B22", d.get("connected_load_kw"),       "#,##0.0")
    fill("D22", d.get("power_factor"),            "0.00")

    # Also update the grid rate assumption in Solar Sizing sheet
    if d.get("rate_per_unit"):
        ws_s = wb["Solar Sizing"]
        ws_s["B11"].value = float(d["rate_per_unit"])
        ws_s["B11"].fill = PatternFill("solid", start_color=YELLOW_INPUT)

    # ── Metadata stamp ────────────────────────────────────────────────────────
    ws_s = wb["Solar Sizing"]
    ws_s["A47"].value = f"Generated by Energybae AI System | {datetime.now().strftime('%d %b %Y, %H:%M')}"
    ws_s["A47"].font = Font(name="Arial", size=8, italic=True, color="888888")

    ws_r = wb["Customer Report"]
    ws_r["B36"].value = f"Report generated: {datetime.now().strftime('%d %B %Y')}"
    ws_r["B36"].font = Font(name="Arial", size=9, italic=True, color="555555")
    ws_r["B36"].alignment = Alignment(horizontal="center")

    wb.save(output_path)
    print(f"  [OK] Excel filled and saved -> {output_path}")
    return output_path


def process_bill(bill_path: str, output_path: str,
                 api_key: str | None = None,
                 template_path: str | None = None,
                 save_json: bool = False) -> dict:
    """
    End-to-end pipeline:
      1. Extract data from bill via Claude AI
      2. Fill Excel template
      3. (Optional) Save extracted JSON
    Returns the extracted data dict.
    """
    print(f"\n{'='*60}")
    print(f"  ENERGYBAE SOLAR LOAD CALCULATOR")
    print(f"{'='*60}")
    print(f"  Bill: {bill_path}")
    print(f"  Output: {output_path}")
    print(f"{'─'*60}")

    print("\n[Step 1/2] Extracting data from electricity bill...")
    data = extract_bill_data(bill_path, api_key)

    if save_json:
        json_path = Path(output_path).with_suffix(".json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"  -> Extracted JSON saved: {json_path}")

    print("\n[Step 2/2] Filling Solar Load Calculator Excel template...")
    fill_excel_template(data, output_path, template_path)

    print(f"\n{'='*60}")
    print(f"  ✅ SUCCESS! Output file ready: {output_path}")
    print(f"{'='*60}\n")
    return data


# ── CLI ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Energybae Solar Load Calculator — Electricity Bill to Excel Automation"
    )
    parser.add_argument("--bill",     required=True, help="Path to electricity bill (PDF/image)")
    parser.add_argument("--output",   default="output/solar_report.xlsx", help="Output Excel file path")
    parser.add_argument("--template", default=None, help="Custom Excel template path (optional)")
    parser.add_argument("--api-key",  default=None, help="Google API key (or set GOOGLE_API_KEY env var)")
    parser.add_argument("--save-json",action="store_true", help="Also save extracted data as JSON")
    args = parser.parse_args()

    if not Path(args.bill).exists():
        print(f"ERROR: Bill file not found: {args.bill}")
        sys.exit(1)

    process_bill(
        bill_path=args.bill,
        output_path=args.output,
        api_key=args.api_key,
        template_path=args.template,
        save_json=args.save_json,
    )


if __name__ == "__main__":
    main()
