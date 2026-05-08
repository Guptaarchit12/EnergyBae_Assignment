"""
Creates the Energybae Solar Load Calculator Excel Template.
This mirrors the standard MSEDCL-based solar sizing sheet used by the sales team.
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import os

# ── Palette ──────────────────────────────────────────────────────────────────
GREEN_DARK   = "1B5E20"
GREEN_MID    = "2E7D32"
GREEN_LIGHT  = "A5D6A7"
GREEN_PALE   = "E8F5E9"
YELLOW_INPUT = "FFF9C4"
ORANGE_CALC  = "FFF3E0"
BLUE_OUTPUT  = "E3F2FD"
HEADER_TXT   = "FFFFFF"
BORDER_CLR   = "BDBDBD"

def side(style="thin", color=BORDER_CLR):
    return Side(style=style, color=color)

def border(top=True, bottom=True, left=True, right=True):
    s = side()
    return Border(
        top=s if top else Side(style=None),
        bottom=s if bottom else Side(style=None),
        left=s if left else Side(style=None),
        right=s if right else Side(style=None),
    )

def cell_style(ws, row, col, value=None, bold=False, bg=None, fg="000000",
               align="left", num_format=None, wrap=False, border_all=True, size=10):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font = Font(name="Arial", bold=bold, color=fg, size=size)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if num_format:
        c.number_format = num_format
    if border_all:
        c.border = border()
    return c


def create_template(output_path: str) -> str:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Customer & Bill Input ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Bill Input"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 26
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 26
    ws1.column_dimensions["E"].width = 18

    # Header banner
    ws1.merge_cells("A1:E1")
    c = ws1["A1"]
    c.value = "⚡ ENERGYBAE — Solar Load Calculator"
    c.font = Font(name="Arial", bold=True, color=HEADER_TXT, size=14)
    c.fill = PatternFill("solid", start_color=GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:E2")
    c = ws1["A2"]
    c.value = "Electricity Bill Data Extraction — MSEDCL / Maharashtra"
    c.font = Font(name="Arial", italic=True, color=GREEN_DARK, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    # Section headers
    def section_header(row, text):
        ws1.merge_cells(f"A{row}:E{row}")
        c = ws1[f"A{row}"]
        c.value = text
        c.font = Font(name="Arial", bold=True, color=HEADER_TXT, size=11)
        c.fill = PatternFill("solid", start_color=GREEN_MID)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws1.row_dimensions[row].height = 22

    # Legend row
    ws1.merge_cells("A3:E3")
    c = ws1["A3"]
    c.value = "🟡 Yellow = Input cells (fill from bill)    🟠 Orange = Auto-calculated    🔵 Blue = Output / Results"
    c.font = Font(name="Arial", size=9, italic=True, color="555555")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[3].height = 18

    # ── Section A: Customer Info ──
    section_header(4, "  A. CUSTOMER INFORMATION")
    fields_a = [
        ("Consumer Name",      "B5",  "", "Consumer Number",    "D5",  ""),
        ("Address",            "B6",  "", "Tariff Category",    "D6",  ""),
        ("Discom / Utility",   "B7",  "MSEDCL", "Division / Sub-div", "D7",  ""),
        ("Bill Month",         "B8",  "", "Bill Date",          "D8",  ""),
        ("Meter Number",       "B9",  "", "Sanctioned Load (kW)","D9", ""),
    ]
    for r_off, (l1, c1, v1, l2, c2, v2) in enumerate(fields_a, start=5):
        row = r_off
        cell_style(ws1, row, 1, l1, bold=True, bg=GREEN_PALE)
        ic = ws1[c1]; ic.value = v1
        ic.fill = PatternFill("solid", start_color=YELLOW_INPUT)
        ic.border = border(); ic.font = Font(name="Arial", size=10)
        ic.alignment = Alignment(horizontal="left", vertical="center")

        cell_style(ws1, row, 3, l2, bold=True, bg=GREEN_PALE)
        ic2 = ws1[c2]; ic2.value = v2
        ic2.fill = PatternFill("solid", start_color=YELLOW_INPUT)
        ic2.border = border(); ic2.font = Font(name="Arial", size=10)
        ic2.alignment = Alignment(horizontal="left", vertical="center")

    # ── Section B: Monthly Consumption ──
    section_header(11, "  B. MONTHLY CONSUMPTION (Units / kWh)")
    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    cell_style(ws1, 12, 1, "Month", bold=True, bg=GREEN_LIGHT, align="center")
    for i, m in enumerate(months, start=2):
        cell_style(ws1, 12, i, m, bold=True, bg=GREEN_LIGHT, align="center")
    cell_style(ws1, 13, 1, "Units (kWh)", bold=True, bg=GREEN_PALE)
    for i in range(2, 14):
        c = ws1.cell(row=13, column=i)
        c.fill = PatternFill("solid", start_color=YELLOW_INPUT)
        c.border = border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(name="Arial", size=10)
    ws1.row_dimensions[12].height = 20
    ws1.row_dimensions[13].height = 22

    # Average & Peak row
    cell_style(ws1, 14, 1, "Average Monthly (kWh)", bold=True, bg=ORANGE_CALC)
    c_avg = ws1.cell(row=14, column=2)
    c_avg.value = "=IFERROR(AVERAGE(B13:M13),\"\")"
    c_avg.fill = PatternFill("solid", start_color=ORANGE_CALC)
    c_avg.border = border(); c_avg.number_format = "#,##0.0"
    c_avg.font = Font(name="Arial", size=10); c_avg.alignment = Alignment(horizontal="center", vertical="center")

    cell_style(ws1, 14, 3, "Peak Month (kWh)", bold=True, bg=ORANGE_CALC)
    c_pk = ws1.cell(row=14, column=4)
    c_pk.value = "=IFERROR(MAX(B13:M13),\"\")"
    c_pk.fill = PatternFill("solid", start_color=ORANGE_CALC)
    c_pk.border = border(); c_pk.number_format = "#,##0.0"
    c_pk.font = Font(name="Arial", size=10); c_pk.alignment = Alignment(horizontal="center", vertical="center")
    cell_style(ws1, 14, 5, "", bg=ORANGE_CALC)

    # ── Section C: Bill Details ──
    section_header(16, "  C. BILL DETAILS & TARIFF")
    bill_fields = [
        ("Total Units Consumed (Latest Bill)", "B17", "", "#,##0",   "Bill Amount (₹)",          "D17", "", "#,##0.00"),
        ("Fixed / Demand Charges (₹)",        "B18", "", "#,##0.00","Electricity Duty (₹)",      "D18", "", "#,##0.00"),
        ("Fuel Adjustment Charge (₹)",        "B19", "", "#,##0.00","Meter Rent / Other (₹)",    "D19", "", "#,##0.00"),
        ("Subsidies / Rebate (₹)",            "B20", "", "#,##0.00","Net Payable Amount (₹)",    "D20", "", "#,##0.00"),
        ("Tariff Slab (e.g. LT-II)",          "B21", "", "@",       "Rate per Unit (₹/kWh)",     "D21", "", "#,##0.00"),
        ("Connected Load (kW)",               "B22", "", "#,##0.0", "Power Factor",               "D22", "", "0.00"),
    ]
    for r_off, (l1, c1, v1, fmt1, l2, c2, v2, fmt2) in enumerate(bill_fields, start=17):
        row = r_off
        cell_style(ws1, row, 1, l1, bold=True, bg=GREEN_PALE)
        ic = ws1[c1]; ic.value = v1
        ic.fill = PatternFill("solid", start_color=YELLOW_INPUT)
        ic.border = border(); ic.font = Font(name="Arial", size=10)
        ic.number_format = fmt1; ic.alignment = Alignment(horizontal="left", vertical="center")

        cell_style(ws1, row, 3, l2, bold=True, bg=GREEN_PALE)
        ic2 = ws1[c2]; ic2.value = v2
        ic2.fill = PatternFill("solid", start_color=YELLOW_INPUT)
        ic2.border = border(); ic2.font = Font(name="Arial", size=10)
        ic2.number_format = fmt2; ic2.alignment = Alignment(horizontal="left", vertical="center")

    # ── Sheet 2: Solar Sizing Calculation ────────────────────────────────────
    ws2 = wb.create_sheet("Solar Sizing")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 30
    ws2.column_dimensions["E"].width = 22

    # Header
    ws2.merge_cells("A1:E1")
    c = ws2["A1"]
    c.value = "⚡ ENERGYBAE — Solar System Sizing Engine"
    c.font = Font(name="Arial", bold=True, color=HEADER_TXT, size=14)
    c.fill = PatternFill("solid", start_color=GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 36

    ws2.merge_cells("A2:E2")
    c2h = ws2["A2"]
    c2h.value = "Auto-calculated from Bill Input sheet — do not manually edit formula cells"
    c2h.font = Font(name="Arial", italic=True, color=GREEN_DARK, size=10)
    c2h.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 18

    def s2_section(row, text):
        ws2.merge_cells(f"A{row}:E{row}")
        c = ws2[f"A{row}"]
        c.value = text
        c.font = Font(name="Arial", bold=True, color=HEADER_TXT, size=11)
        c.fill = PatternFill("solid", start_color=GREEN_MID)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.row_dimensions[row].height = 22

    def s2_row(row, label, formula_or_val, unit="", note="", is_output=False):
        bg = BLUE_OUTPUT if is_output else ORANGE_CALC
        cell_style(ws2, row, 1, label, bold=is_output, bg=GREEN_PALE)
        vc = ws2.cell(row=row, column=2)
        vc.value = formula_or_val
        vc.fill = PatternFill("solid", start_color=bg)
        vc.border = border(); vc.font = Font(name="Arial", size=10, bold=is_output)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        cell_style(ws2, row, 3, unit, bg=bg)
        cell_style(ws2, row, 4, note, bg="F5F5F5", size=9)
        ws2.row_dimensions[row].height = 20

    def s2_assumption(row, label, val, unit="", note=""):
        cell_style(ws2, row, 1, label, bold=True, bg=GREEN_PALE)
        vc = ws2.cell(row=row, column=2)
        vc.value = val
        vc.fill = PatternFill("solid", start_color=YELLOW_INPUT)
        vc.border = border(); vc.font = Font(name="Arial", size=10, color="0000FF")
        vc.alignment = Alignment(horizontal="right", vertical="center")
        cell_style(ws2, row, 3, unit, bg=YELLOW_INPUT)
        cell_style(ws2, row, 4, note, bg="F5F5F5", size=9)
        ws2.row_dimensions[row].height = 20

    # ── Section 1: Assumptions ──
    s2_section(3, "  1. SYSTEM ASSUMPTIONS (Editable)")
    s2_assumption(4,  "Peak Sun Hours (PSH) — Maharashtra",  4.5,  "hrs/day",  "Avg solar irradiance for Pune/MH region")
    s2_assumption(5,  "System Efficiency",                   0.80, "",         "Accounts for inverter, wiring, soiling losses")
    s2_assumption(6,  "Performance Ratio",                   0.75, "",         "Typical PR for rooftop systems in India")
    s2_assumption(7,  "Cost per kWp (On-Grid)",              45000,"₹/kWp",    "Market rate as of 2024-25; adjust as needed")
    s2_assumption(8,  "Cost per kWp (Off-Grid/Hybrid)",      65000,"₹/kWp",    "Includes battery storage")
    s2_assumption(9,  "Annual Degradation Rate",             0.007,"",         "0.7% per year — standard poly/mono panels")
    s2_assumption(10, "Net Metering Applicable?",            "Yes","",         "Yes/No — affects savings calculation")
    s2_assumption(11, "Grid Electricity Rate (₹/kWh)",       9.0,  "₹/kWh",    "Pull from Bill Input D21 or enter manually")
    s2_assumption(12, "Annual Electricity Escalation",       0.05, "",         "5% YoY tariff increase assumption")
    s2_assumption(13, "System Lifespan",                     25,   "years",    "Standard warranty period")
    s2_assumption(14, "Subsidy (PM Surya Ghar scheme)",      0.30, "",         "30% central govt subsidy on system cost")

    # ── Section 2: Demand Analysis ──
    s2_section(16, "  2. ENERGY DEMAND ANALYSIS")
    s2_row(17, "Average Monthly Consumption (kWh)",
           "='Bill Input'!B14", "kWh/month", "Pulled from Bill Input — average of 12 months")
    s2_row(18, "Peak Monthly Consumption (kWh)",
           "='Bill Input'!D14", "kWh/month", "Max month — used for sizing to avoid shortfall")
    s2_row(19, "Daily Energy Requirement (avg)",
           "=IFERROR(B17/30,\"\")", "kWh/day",  "B17 / 30 days")
    s2_row(20, "Daily Energy Requirement (peak)",
           "=IFERROR(B18/30,\"\")", "kWh/day",  "B18 / 30 days — conservative sizing basis")
    s2_row(21, "Annual Energy Consumption",
           "=IFERROR(B17*12,\"\")", "kWh/year", "Annualised from average monthly")

    # ── Section 3: System Sizing ──
    s2_section(23, "  3. RECOMMENDED SOLAR SYSTEM SIZE")
    s2_row(24, "Solar Capacity Required (avg basis)",
           '=IFERROR(B19/(B4*B5),"")' , "kWp",
           "Daily kWh ÷ (PSH × System Efficiency)")
    s2_row(25, "Solar Capacity Required (peak basis)",
           '=IFERROR(B20/(B4*B5),"")' , "kWp",
           "Peak daily kWh ÷ (PSH × System Efficiency) — recommended")
    s2_row(26, "Recommended System Size (rounded)",
           '=IFERROR(CEILING(B25,1),"")' , "kWp",
           "Rounded up to nearest 1 kWp", is_output=True)
    s2_row(27, "Number of Solar Panels (400W)",
           '=IFERROR(CEILING(B26*1000/400,1),"")' , "panels",
           "Assuming 400W monocrystalline panels")
    s2_row(28, "Rooftop Area Required",
           '=IFERROR(B26*10,"")' , "sq. meters",
           "~10 sq.m per kWp rule of thumb")
    s2_row(29, "Annual Solar Generation (Year 1)",
           '=IFERROR(B26*B4*365*B6,"")' , "kWh/year",
           "Capacity × PSH × 365 × Performance Ratio")
    s2_row(30, "Solar Coverage Ratio",
           '=IFERROR(B29/B21,"")' , "",
           "Solar generation ÷ Annual consumption (aim >0.9)")

    # ── Section 4: Financial Analysis ──
    s2_section(32, "  4. FINANCIAL ANALYSIS")
    s2_row(33, "System Cost (On-Grid, before subsidy)",
           '=IFERROR(B26*B7,"")' , "₹",
           "Recommended size × cost per kWp")
    s2_row(34, "Subsidy Amount",
           '=IFERROR(B33*B14,"")' , "₹",
           "PM Surya Ghar or applicable state subsidy")
    s2_row(35, "Net System Cost (after subsidy)",
           '=IFERROR(B33-B34,"")' , "₹",
           "Customer's actual investment", is_output=True)
    s2_row(36, "Annual Electricity Bill (current)",
           "='Bill Input'!D20*12" , "₹/year",
           "Net Payable (monthly) × 12")
    s2_row(37, "Annual Savings — Year 1",
           '=IFERROR(B29*B11,"")' , "₹/year",
           "Solar generation × grid rate")
    s2_row(38, "Simple Payback Period",
           '=IFERROR(B35/B37,"")' , "years",
           "Net cost ÷ Year-1 savings", is_output=True)
    s2_row(39, "25-Year Cumulative Savings",
           '=IFERROR(B37*((1-(1+B12)^(-B13))/B12)*(1+B12),"")' , "₹",
           "NPV of escalating savings over lifespan", is_output=True)
    s2_row(40, "ROI (25 years)",
           '=IFERROR((B39-B35)/B35,"")' , "",
           "Total return on investment", is_output=True)

    # ── Section 5: CO2 Impact ──
    s2_section(42, "  5. ENVIRONMENTAL IMPACT")
    s2_row(43, "CO₂ Avoided per Year",
           '=IFERROR(B29*0.82,"")' , "kg CO₂/year",
           "India grid emission factor: 0.82 kg CO₂/kWh (CEA 2023)")
    s2_row(44, "Trees Equivalent per Year",
           '=IFERROR(B43/21,"")' , "trees",
           "~21 kg CO₂ absorbed per tree per year")
    s2_row(45, "Lifetime CO₂ Avoided",
           '=IFERROR(B43*B13,"")' , "kg CO₂",
           "Over 25-year system lifespan")

    # ── Sheet 3: Summary Report ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Customer Report")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 3
    ws3.column_dimensions["B"].width = 36
    ws3.column_dimensions["C"].width = 26
    ws3.column_dimensions["D"].width = 4

    # Header
    ws3.merge_cells("B1:C1")
    c = ws3["B1"]
    c.value = "ENERGYBAE — Solar Proposal Summary"
    c.font = Font(name="Arial", bold=True, color=HEADER_TXT, size=16)
    c.fill = PatternFill("solid", start_color=GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 44

    ws3.merge_cells("B2:C2")
    c = ws3["B2"]
    c.value = "www.energybae.in  |  +91 9112233120  |  energybae.co@gmail.com"
    c.font = Font(name="Arial", size=9, italic=True, color=GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[2].height = 18

    def report_row(row, label, formula, bold_val=False, num_fmt=None, is_highlight=False):
        bg = GREEN_PALE
        c_l = ws3.cell(row=row, column=2)
        c_l.value = label
        c_l.font = Font(name="Arial", size=11, bold=True)
        c_l.fill = PatternFill("solid", start_color=bg)
        c_l.border = border(); c_l.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        c_v = ws3.cell(row=row, column=3)
        c_v.value = formula
        c_v.font = Font(name="Arial", size=11, bold=bold_val,
                        color=GREEN_DARK if is_highlight else "000000")
        c_v.fill = PatternFill("solid", start_color=BLUE_OUTPUT if is_highlight else "FFFFFF")
        c_v.border = border(); c_v.alignment = Alignment(horizontal="center", vertical="center")
        if num_fmt:
            c_v.number_format = num_fmt
        ws3.row_dimensions[row].height = 22

    def report_section(row, text):
        ws3.merge_cells(f"B{row}:C{row}")
        c = ws3[f"B{row}"]
        c.value = text
        c.font = Font(name="Arial", bold=True, color=HEADER_TXT, size=11)
        c.fill = PatternFill("solid", start_color=GREEN_MID)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws3.row_dimensions[row].height = 24

    report_section(4, "  CUSTOMER DETAILS")
    report_row(5, "Customer Name",      "='Bill Input'!B5")
    report_row(6, "Consumer Number",    "='Bill Input'!D5")
    report_row(7, "Address",            "='Bill Input'!B6")
    report_row(8, "Tariff Category",    "='Bill Input'!D6")

    report_section(10, "  CONSUMPTION SUMMARY")
    report_row(11, "Average Monthly Usage",  "='Solar Sizing'!B17", num_fmt="#,##0.0 \"kWh\"")
    report_row(12, "Peak Monthly Usage",     "='Solar Sizing'!B18", num_fmt="#,##0.0 \"kWh\"")
    report_row(13, "Current Annual Bill",    "='Solar Sizing'!B36", num_fmt="\"₹\"#,##0")

    report_section(15, "  RECOMMENDED SOLAR SOLUTION")
    report_row(16, "Recommended System Size", "='Solar Sizing'!B26", bold_val=True, num_fmt="0.0\" kWp\"", is_highlight=True)
    report_row(17, "Number of Panels (400W)", "='Solar Sizing'!B27", num_fmt="0\" panels\"")
    report_row(18, "Rooftop Area Required",   "='Solar Sizing'!B28", num_fmt="0\" sq.m\"")
    report_row(19, "Annual Generation",       "='Solar Sizing'!B29", num_fmt="#,##0.0\" kWh/yr\"")

    report_section(21, "  FINANCIAL SUMMARY")
    report_row(22, "Total System Cost",       "='Solar Sizing'!B33", num_fmt="\"₹\"#,##0")
    report_row(23, "Govt. Subsidy",           "='Solar Sizing'!B34", num_fmt="\"₹\"#,##0")
    report_row(24, "Net Investment",          "='Solar Sizing'!B35", bold_val=True, num_fmt="\"₹\"#,##0", is_highlight=True)
    report_row(25, "Annual Savings (Year 1)", "='Solar Sizing'!B37", num_fmt="\"₹\"#,##0")
    report_row(26, "Payback Period",          "='Solar Sizing'!B38", bold_val=True, num_fmt="0.0\" years\"", is_highlight=True)
    report_row(27, "25-Year Savings",         "='Solar Sizing'!B39", bold_val=True, num_fmt="\"₹\"#,##0", is_highlight=True)
    report_row(28, "Return on Investment",    "='Solar Sizing'!B40", num_fmt="0.0%")

    report_section(30, "  ENVIRONMENTAL IMPACT")
    report_row(31, "CO₂ Avoided per Year", "='Solar Sizing'!B43", num_fmt="#,##0\" kg CO₂/yr\"")
    report_row(32, "Equivalent Trees Planted", "='Solar Sizing'!B44", num_fmt="#,##0\" trees\"")

    # Footer
    ws3.merge_cells("B34:C34")
    c = ws3["B34"]
    c.value = "This is an AI-generated estimate. Final sizing subject to site survey. Valid 30 days from date of report."
    c.font = Font(name="Arial", size=8, italic=True, color="888888")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws3.row_dimensions[34].height = 28

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    out = create_template("templates/solar_load_calculator_template.xlsx")
    print(f"Template created: {out}")
